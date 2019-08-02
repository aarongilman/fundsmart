# portfolios app admin.py
import xlrd
import openpyxl
import logging

from django import forms
from django.urls import path
from django.urls import reverse
from django.contrib import admin
from django.shortcuts import render, redirect
from django.utils.translation import ugettext_lazy as _

from accounts.models import User
from .models import Security, Price, FXRate, FundHolding, FundDetail

logger = logging.getLogger('fundsmart')


class CsvImportForm(forms.Form):
    data_file = forms.FileField()


@admin.register(Security)
class SecurityAdmin(admin.ModelAdmin):

    change_list_template = "portfolios/custom_changelist.html"
    list_display = ('name', 'isin', 'date', 'asset_type', 'industry')

    def get_urls(self):
        urls = super().get_urls()
        return [path('import-csv/', self.import_csv)] + urls

    def import_csv(self, request):
        if request.method == "POST":
            user = User.objects.get(username=request.user)
            data_file = request.FILES["data_file"]
            wb = openpyxl.load_workbook(data_file, data_only=True)
            worksheet = wb.active
            objects = []
            try:
                for row in worksheet.iter_rows(min_row=2):
                    objects.append(
                        Security(date=row[0].value, name=row[1].value,
                                 id_value=row[2].value, isin=row[3].value,
                                 ticker=row[4].value, asset_type=row[5].value,
                                 currency=row[6].value, country=row[7].value,
                                 industry=row[8].value, rating=row[9].value,
                                 created_by=user))
                Security.objects.bulk_create(objects, ignore_conflicts=True)
                self.message_user(request, _("Your file has been imported"))
            except Exception as e:
                logger.error(
                    'Error {} occurred while uploading securities'.format(e))
                self.message_user(request, _("Failed to import file"))
            return redirect(reverse('admin:portfolios_security_changelist'))
        form = CsvImportForm()
        return render(request, "portfolios/csv_form.html", {'form': form})


def chunks(l, n):
    # For item i in a range that is a length of l,
    for i in range(0, len(l), n):
        # Create an index range for l of n items:
        yield l[i:i+n]


@admin.register(Price)
class PriceAdmin(admin.ModelAdmin):

    change_list_template = "portfolios/custom_changelist.html"
    list_display = ('date', 'id_value', 'price')

    def get_urls(self):
        urls = super().get_urls()
        return [path('import-csv/', self.import_csv)] + urls

    def import_csv(self, request):
        if request.method == "POST":
            user = User.objects.get(username=request.user)
            data_file = request.FILES["data_file"]
            wb = openpyxl.load_workbook(data_file, data_only=True, read_only=True)
            worksheet = wb.active
            objects = []
            try:
                for row in worksheet.iter_rows(min_row=2):
                    objects.append(
                        Price(date=row[0].value, id_value=row[1].value,
                              price=row[2].value, created_by=user))
                objects = list(chunks(objects, 50000))
                for obj in objects:
                    Price.objects.bulk_create(obj, ignore_conflicts=True)
                self.message_user(request, _("Your file has been imported"))
            except Exception as e:
                logger.error(
                    'Error {} occurred while uploading Price data'.format(e))
                self.message_user(request, _("Failed to import file"))

            return redirect(reverse('admin:portfolios_price_changelist'))
        form = CsvImportForm()
        return render(request, "portfolios/csv_form.html", {'form': form})


@admin.register(FXRate)
class FXRateAdmin(admin.ModelAdmin):

    change_list_template = "portfolios/custom_changelist.html"
    list_display = ('date', 'currency', 'base', 'rate')

    def get_urls(self):
        urls = super().get_urls()
        return [path('import-csv/', self.import_csv)] + urls

    def import_csv(self, request):
        if request.method == "POST":
            user = User.objects.get(username=request.user)
            data_file = request.FILES["data_file"]
            wb = openpyxl.load_workbook(data_file, data_only=True)
            worksheet = wb.active
            objects = []
            try:
                for row in worksheet.iter_rows(min_row=2):
                    objects.append(
                        FXRate(date=row[0].value, currency=row[1].value,
                               base=row[2].value, rate=row[3].value,
                               created_by=user))
                FXRate.objects.bulk_create(objects, ignore_conflicts=True)
                self.message_user(request, _("Your file has been imported"))
            except Exception as e:
                logger.error(
                    'Error {} occurred while uploading FXrate'.format(e))
                self.message_user(request, _("Failed to import file"))

            return redirect(reverse('admin:portfolios_fxrate_changelist'))
        form = CsvImportForm()
        return render(request, "portfolios/csv_form.html", {'form': form})


@admin.register(FundHolding)
class FundHoldingAdmin(admin.ModelAdmin):

    change_list_template = "portfolios/custom_changelist.html"
    list_display = ('date', 'fund', 'fund_id', 'id_value', 'quantity',
                    'market_value', 'net_asset_percentage')

    def get_urls(self):
        urls = super().get_urls()
        return [path('import-csv/', self.import_csv)] + urls

    def import_csv(self, request):
        if request.method == "POST":
            user = User.objects.get(username=request.user)
            data_file = request.FILES["data_file"]
            wb = openpyxl.load_workbook(data_file, data_only=True)
            worksheet = wb.active
            objects = []
            try:
                for row in worksheet.iter_rows(min_row=2):
                    objects.append(
                        FundHolding(date=row[0].value, fund=row[1].value,
                                    fund_id=row[2].value, id_value=row[3].value,
                                    quantity=row[4].value,
                                    market_value=float(row[5].value) if row[5].value else None,
                                    net_asset_percentage=float(row[6].value) if row[6].value else None,
                                    created_by=user))
                FundHolding.objects.bulk_create(objects, ignore_conflicts=True)
                self.message_user(request, _("Your file has been imported"))
            except Exception as e:
                logger.error(
                    'Error {} occurred while uploading fund holding'.format(e))
                self.message_user(request, _("Failed to import file"))

            return redirect(reverse('admin:portfolios_fundholding_changelist'))
        form = CsvImportForm()
        return render(request, "portfolios/csv_form.html", {'form': form})


@admin.register(FundDetail)
class FundDetailAdmin(admin.ModelAdmin):

    change_list_template = "portfolios/custom_changelist.html"
    list_display = ('asset_type', 'category', 'name', 'fund_id', 'benchmark',
                    'aum', 'regular', 'direct', 'fund_exp_ratio')

    def get_urls(self):
        urls = super().get_urls()
        return [path('import-csv/', self.import_csv)] + urls

    def import_csv(self, request):
        if request.method == "POST":
            user = User.objects.get(username=request.user)
            data_file = request.FILES["data_file"]
            wb = xlrd.open_workbook(file_contents=data_file.read())
            ws = wb.sheet_by_index(0)

            objects = []
            try:
                for i in range(2, ws.nrows):
                    row = ws.row_values(i)
                    objects.append(
                        FundDetail(asset_type=row[0], category=row[1],
                                   name=row[2], fund_id=row[3],
                                   benchmark=row[4],
                                   aum=row[5] if row[5] else None,
                                   regular=row[6], direct=row[7],
                                   fund_exp_ratio=row[8],
                                   return_1_year=row[9] if row[9] else None,
                                   return_3_year=row[10] if row[10] else None,
                                   return_5_year=row[11] if row[11] else None,
                                   benchmark_1_year=row[12] if row[12] else None,
                                   benchmark_3_year=row[13] if row[13] else None,
                                   benchmark_5_year=row[14] if row[14] else None,
                                   return_over_bench_1_year=row[15] if row[15] else None,
                                   return_over_bench_3_year=row[16] if row[16] else None,
                                   return_over_bench_5_year=row[17] if row[17] else None,
                                   for_recommendation=row[18] if row[18] else None,
                                   created_by=user))
                FundDetail.objects.bulk_create(objects, ignore_conflicts=True)
                self.message_user(request, _("Your file has been imported"))
            except Exception as e:
                logger.error(
                    'Error {} occurred while uploading fund details'.format(e))
                self.message_user(request, _("Failed to import file"))

            return redirect(reverse('admin:portfolios_funddetail_changelist'))
        form = CsvImportForm()
        return render(request, "portfolios/csv_form.html", {'form': form})
